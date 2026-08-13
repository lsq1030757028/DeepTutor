"""save_delivery 的离线测试：多格式落盘（含 Postman Collection）+ 收据校验。"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re

import pytest

from server import delivery

CASES = [
    {
        "编号": "TC-001",
        "标题": "已登录用户分页查询订单列表成功",
        "前置条件": "持有效 token",
        "操作步骤": ["调用 GET /api/v1/orders，page=1", "检查响应结构"],
        "预期结果": "状态码 200；code=0；data.total 为整数",
        "优先级": "高",
        "所属模块": "订单",
        "用例类型": "功能测试",
        "测试数据": "page=1",
        "关联端点": ["GET /api/v1/orders"],
        "request": {
            "method": "GET",
            "url": "https://api.shop.example.com/api/v1/orders?page=1",
            "headers": [{"key": "Authorization", "value": "{{token}}"}],
            "body": {"mode": "none"},
            "assertions": [{"type": "status", "expected": 200},
                           {"type": "json_path", "path": "$.data.total", "expected": 2}],
        },
    },
    {
        "编号": "TC-002",
        "标题": "订单状态不允许退款时退款被拒",
        "前置条件": "存在一笔已完结订单",
        "操作步骤": ["调用 POST /api/v1/orders/:id/refund"],
        "预期结果": "状态码 400；code=40001；提示订单状态不允许退款",
        "优先级": "中",
        "关联端点": ["POST /api/v1/orders/:id/refund"],
    },
]


@pytest.fixture
def out_root(tmp_path, monkeypatch):
    root = tmp_path / "deliveries"
    monkeypatch.setattr(delivery, "DELIVERIES_DIR", str(root))
    return root


def _receipt(result):
    with open(os.path.join(result["delivery_dir"], "receipt.json"), encoding="utf-8") as f:
        return json.load(f)


# ── 目录与命名 ──────────────────────────────────────────────────────────────

def test_directory_naming(out_root):
    result = delivery.save_delivery(CASES, title="订单域 HAR 用例", fmt="csv")
    assert result["ok"] is True
    name = os.path.basename(result["delivery_dir"])
    assert re.match(r"^\d{8}-\d{6}-", name), name
    assert "订单域" in name
    assert os.path.isdir(result["delivery_dir"])


def test_default_format_is_xlsx_plus_postman(out_root):
    """M2.5：HAR 链路默认双产物——xlsx 给人读/进 TAPD，collection 给机器跑。"""
    result = delivery.save_delivery(CASES, title="默认格式")
    assert result["format"] == "xlsx+postman"
    assert delivery.DEFAULT_FORMAT == "xlsx+postman"
    names = [os.path.basename(p) for p in result["files"]]
    # cases.json 是 0.7.0 起无条件落的边车文件（工作台按它列用例、勾选后执行）
    assert names == ["cases.xlsx", "默认格式.postman_collection.json",
                     "cases.json", "receipt.json"]
    assert all(os.path.isfile(p) for p in result["files"])
    assert "Apifox" in result["import_hint"]


# ── xlsx ────────────────────────────────────────────────────────────────────

def test_xlsx_columns_align_tapd(out_root):
    from openpyxl import load_workbook

    result = delivery.save_delivery(CASES, title="订单用例", fmt="xlsx")
    wb = load_workbook(result["files"][0])
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == list(delivery.TAPD_COLUMNS)
    assert ws.max_row == 3                       # 表头 + 2 条用例
    assert ws.cell(row=2, column=1).value == "已登录用户分页查询订单列表成功"
    assert ws.cell(row=2, column=6).value == "高"          # 用例等级
    assert "1. " in ws.cell(row=2, column=4).value        # 多步骤自动编号
    assert ws.cell(row=3, column=2).value == delivery.DEFAULT_MODULE
    assert ws.cell(row=3, column=7).value == delivery.DEFAULT_CASE_TYPE


# ── csv ─────────────────────────────────────────────────────────────────────

def test_csv_roundtrip(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="csv")
    path = result["files"][0]
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    assert list(rows[0].keys()) == list(delivery.TAPD_COLUMNS)
    assert len(rows) == 2
    assert rows[1]["用例等级"] == "中"
    with open(path, "rb") as f:
        assert f.read(3) == b"\xef\xbb\xbf"      # BOM：Excel 双击不乱码


# ── markdown ────────────────────────────────────────────────────────────────

def test_markdown_has_id_column_and_metadata(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="markdown",
                                    source_fingerprint="sha256:abc123")
    text = open(result["files"][0], encoding="utf-8").read()
    assert text.startswith("# 订单用例")
    assert "sha256:abc123" in text
    assert "| 编号 | " + " | ".join(delivery.TAPD_COLUMNS) + " |" in text
    assert "TC-001" in text and "TC-002" in text
    assert "<br>" in text                        # 多步骤换行转义，表格不散


def test_format_aliases(out_root):
    assert delivery.save_delivery(CASES, title="别名", fmt="md")["format"] == "markdown"
    assert delivery.save_delivery(CASES, title="别名", fmt="Excel")["format"] == "xlsx"
    assert delivery.save_delivery(CASES, title="别名",
                                  fmt="Collection")["format"] == "postman"
    assert delivery.save_delivery(CASES, title="别名",
                                  fmt="postman+xlsx")["format"] == "xlsx+postman"


# ── postman collection ──────────────────────────────────────────────────────

def _collection(result):
    path = next(p for p in result["files"] if p.endswith(".postman_collection.json"))
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def test_postman_only_writes_one_named_product(out_root):
    result = delivery.save_delivery(CASES, title="订单域接口用例", fmt="postman")
    names = [os.path.basename(p) for p in result["files"]]
    assert names == ["订单域接口用例.postman_collection.json",
                     "cases.json", "receipt.json"]


def test_postman_collection_schema_and_shape(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="postman",
                                    source_fingerprint="sha256:abc123")
    col = _collection(result)
    assert col["info"]["schema"] == \
        "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
    assert col["info"]["name"] == "订单用例"
    assert col["info"]["_postman_id"]
    assert "sha256:abc123" in col["info"]["description"]
    # item 按所属模块分文件夹：订单 + 未分类（TC-002 没写模块）
    folders = {f["name"]: f for f in col["item"]}
    assert set(folders) == {"订单", delivery.DEFAULT_MODULE}
    assert sum(len(f["item"]) for f in col["item"]) == 2
    assert {v["key"] for v in col["variable"]} == {"baseUrl"}


def test_postman_item_request_and_test_script(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="postman")
    item = _collection(result)["item"][0]["item"][0]
    assert item["name"] == "TC-001 已登录用户分页查询订单列表成功"
    req = item["request"]
    assert req["method"] == "GET"
    assert req["header"] == [{"key": "Authorization", "value": "{{token}}", "type": "text"}]
    assert req["url"]["raw"] == "https://api.shop.example.com/api/v1/orders?page=1"
    assert req["url"]["host"] == ["api", "shop", "example", "com"]
    assert req["url"]["path"] == ["api", "v1", "orders"]
    assert req["url"]["query"] == [{"key": "page", "value": "1"}]
    assert "前置条件：持有效 token" in req["description"]
    script = "\n".join(item["event"][0]["script"]["exec"])
    assert item["event"][0]["listen"] == "test"
    assert "pm.response.to.have.status(200);" in script
    assert 'pm.expect(jsonData["data"]["total"]).to.eql(2);' in script


def test_postman_base_url_variable_picks_the_absolute_origin(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="postman")
    variable = _collection(result)["variable"][0]
    assert variable["value"] == "https://api.shop.example.com"


def test_postman_placeholder_item_for_case_without_request(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="postman")
    plain = _collection(result)["item"][1]["item"][0]      # TC-002 没有 request 块
    assert plain["name"].startswith("TC-002")
    assert "人执行用例，无结构化请求" in plain["request"]["description"]
    assert "event" not in plain
    assert result["postman"]["placeholder_count"] == 1
    assert any("没有 request 块" in w for w in result["warnings"])


def test_postman_body_carries_language_option(out_root):
    cases = [dict(CASES[0], request=dict(
        CASES[0]["request"], method="POST",
        url="{{baseUrl}}/api/v1/orders",
        body={"mode": "raw", "language": "json", "raw": '{"skuId":"SKU-1","qty":1}'}))]
    result = delivery.save_delivery(cases, title="下单", fmt="postman")
    body = _collection(result)["item"][0]["item"][0]["request"]["body"]
    assert body == {"mode": "raw", "raw": '{"skuId":"SKU-1","qty":1}',
                    "options": {"raw": {"language": "json"}}}


def test_postman_warns_when_base_url_unresolved(out_root):
    cases = [dict(CASES[0], request=dict(CASES[0]["request"],
                                         url="{{baseUrl}}/api/v1/orders?page=1"))]
    result = delivery.save_delivery(cases, title="订单用例", fmt="postman")
    assert _collection(result)["variable"][0]["value"] == ""
    assert any("baseUrl" in w for w in result["warnings"])


def test_postman_json_is_utf8_and_reparsable(out_root):
    """中文不转义、文件可被 json 解析——Apifox 导入的前提。"""
    result = delivery.save_delivery(CASES, title="订单用例", fmt="postman")
    path = result["postman_file"]
    text = open(path, encoding="utf-8").read()
    assert "已登录用户分页查询订单列表成功" in text          # 未被 \\uXXXX 转义
    assert json.loads(text)["info"]["name"] == "订单用例"


# ── 双产物 ──────────────────────────────────────────────────────────────────

def test_dual_format_writes_both_and_records_both_hashes(out_root):
    from openpyxl import load_workbook

    result = delivery.save_delivery(CASES, title="订单用例", fmt="xlsx+postman",
                                    source_fingerprint="sha256:deadbeef")
    assert result["ok"] is True and result["format"] == "xlsx+postman"
    receipt = _receipt(result)
    assert [a["file"] for a in receipt["artifacts"]] == [
        "cases.xlsx", "订单用例.postman_collection.json"]
    for art in receipt["artifacts"]:
        blob = open(os.path.join(result["delivery_dir"], art["file"]), "rb").read()
        assert art["sha256"] == hashlib.sha256(blob).hexdigest()
        assert art["bytes"] == len(blob)
    assert receipt["postman"]["item_count"] == 2
    assert receipt["postman"]["assertion_count"] == 2
    assert receipt["postman"]["schema"].endswith("collection.json")
    # 两份产物同源：xlsx 的行数与 collection 的 item 数对得上
    ws = load_workbook(os.path.join(result["delivery_dir"], "cases.xlsx")).active
    assert ws.max_row - 1 == sum(len(f["item"]) for f in _collection(result)["item"])


# ── 收据 ────────────────────────────────────────────────────────────────────

def test_receipt_records_hashes_and_validation(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="csv",
                                    source_fingerprint="sha256:deadbeef")
    receipt = _receipt(result)
    assert receipt["schema"] == delivery.RECEIPT_SCHEMA
    assert receipt["format"] == "csv"
    assert receipt["case_count"] == 2
    assert receipt["source_fingerprint"] == "sha256:deadbeef"
    assert receipt["input_fingerprint"].startswith("sha256:")
    assert receipt["validation"]["ok"] is True
    assert receipt["validation"]["error_count"] == 0
    assert receipt["columns"] == list(delivery.TAPD_COLUMNS)
    assert receipt["case_index"] == [
        {"编号": "TC-001", "用例名称": CASES[0]["标题"]},
        {"编号": "TC-002", "用例名称": CASES[1]["标题"]},
    ]
    art = receipt["artifacts"][0]
    product = os.path.join(result["delivery_dir"], art["file"])
    with open(product, "rb") as f:
        blob = f.read()
    assert art["sha256"] == hashlib.sha256(blob).hexdigest()
    assert art["bytes"] == len(blob)


def test_input_fingerprint_is_stable_and_content_bound(out_root):
    a = _receipt(delivery.save_delivery(CASES, title="a", fmt="csv"))
    b = _receipt(delivery.save_delivery(list(CASES), title="b", fmt="csv"))
    changed = [dict(CASES[0], 优先级="低"), CASES[1]]
    c = _receipt(delivery.save_delivery(changed, title="c", fmt="csv"))
    assert a["input_fingerprint"] == b["input_fingerprint"]
    assert a["input_fingerprint"] != c["input_fingerprint"]


def test_receipt_flags_invalid_cases(out_root):
    bad = [dict(CASES[0], 优先级="P0")]
    result = delivery.save_delivery(bad, title="有问题的用例", fmt="csv")
    assert result["ok"] is True                  # 仍然落盘
    assert result["validation_ok"] is False      # 但收据据实记录
    receipt = _receipt(result)
    assert receipt["validation"]["ok"] is False
    assert receipt["validation"]["errors"][0]["code"] == "E08_PRIORITY_INVALID"


# ── 错误路径 ────────────────────────────────────────────────────────────────

def test_unsupported_format(out_root):
    result = delivery.save_delivery(CASES, title="x", fmt="xmind")
    assert result["ok"] is False and result["error"] == "FORMAT_UNSUPPORTED"
    assert result["hint"]


def test_empty_cases(out_root):
    result = delivery.save_delivery([], title="x", fmt="csv")
    assert result["ok"] is False and result["error"] == "CASES_EMPTY"


def test_all_invalid_shape(out_root):
    result = delivery.save_delivery(["纯字符串"], title="x", fmt="csv")
    assert result["ok"] is False and result["error"] == "CASES_ALL_INVALID"


def test_title_slug_falls_back(out_root):
    result = delivery.save_delivery(CASES, title="!!!", fmt="csv")
    assert os.path.basename(result["delivery_dir"]).endswith("-untitled")


def test_same_second_same_title_does_not_overwrite(out_root):
    first = delivery.save_delivery(CASES, title="同名交付", fmt="csv")
    second = delivery.save_delivery(CASES, title="同名交付", fmt="csv")
    assert first["delivery_dir"] != second["delivery_dir"]
    assert os.path.isfile(os.path.join(first["delivery_dir"], "cases.csv"))
    assert os.path.isfile(os.path.join(second["delivery_dir"], "cases.csv"))


# ── 入参宽容（热修 0.3.1：修复优于拒绝） ────────────────────────────────────

def test_cases_as_json_string_saved_and_echoed(out_root):
    result = delivery.save_delivery(json.dumps(CASES, ensure_ascii=False),
                                    title="字符串入参", fmt="csv")
    assert result["ok"] is True and result["case_count"] == 2
    assert any("JSON" in n for n in result["normalized"])
    assert _receipt(result)["normalized"] == result["normalized"]


def test_single_case_object_saved(out_root):
    result = delivery.save_delivery(CASES[0], title="单条用例", fmt="csv")
    assert result["ok"] is True and result["case_count"] == 1
    assert any("单元素数组" in n for n in result["normalized"])


def test_envelope_carries_format_and_title(out_root):
    """整套实参被包进 cases：format/title 得从信封里捡回来，不能退回默认值。"""
    result = delivery.save_delivery({"cases": CASES, "format": " CSV ",
                                     "title": "信封交付",
                                     "source_fingerprint": "sha256:abc123"})
    assert result["ok"] is True and result["format"] == "csv"
    assert "信封交付" in os.path.basename(result["delivery_dir"])
    assert _receipt(result)["source_fingerprint"] == "sha256:abc123"


def test_nested_wrapper_peeled(out_root):
    result = delivery.save_delivery({"input": {"arguments": {"cases": CASES}}},
                                    title="三层包裹", fmt="csv")
    assert result["ok"] is True and result["case_count"] == 2


def test_broken_json_string_refused_with_hint(out_root):
    result = delivery.save_delivery('[{"编号": "TC-1"', title="坏 JSON", fmt="csv")
    assert result["ok"] is False and result["error"] == "CASES_JSON_INVALID"
    assert "不是合法 JSON" in result["message"] and result["hint"]


def test_well_formed_call_has_no_normalized_key(out_root):
    result = delivery.save_delivery(CASES, title="正常入参", fmt="csv")
    assert "normalized" not in result and "normalized" not in _receipt(result)


# ── cases.json（0.7.0 的机器可读边车） ──────────────────────────────────────
# xlsx 丢了 request 块、collection 把断言编译成了 JS——想把一批用例原样读回来
# （工作台就要干这个）只能靠这一份。所以它无条件落盘，且形状要能直接喂回执行器。

def _cases_json(result):
    with open(os.path.join(result["delivery_dir"], "cases.json"), encoding="utf-8") as f:
        return json.load(f)


def test_cases_json_is_written_for_every_format(out_root):
    for fmt in ("xlsx", "csv", "markdown", "postman", "xlsx+postman"):
        result = delivery.save_delivery(CASES, title="边车" + fmt, fmt=fmt)
        assert os.path.isfile(result["cases_file"]), fmt
        assert _cases_json(result)["case_count"] == 2, fmt


def test_cases_json_shape(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="csv",
                                    source_fingerprint="sha256:abc123")
    payload = _cases_json(result)
    assert payload["schema"] == delivery.CASES_SCHEMA
    assert payload["title"] == "订单用例"
    assert payload["source_fingerprint"] == "sha256:abc123"
    assert payload["generated_at"]
    first = payload["cases"][0]
    assert first["case_id"] == "TC-001"
    assert first["title"] == CASES[0]["标题"]
    assert first["priority"] == "高"
    assert first["module"] == "订单"
    assert isinstance(first["steps"], list) and first["steps"]
    # 内部标记不许落盘（_present 还是个 set，本来也序列化不了）
    assert not [k for k in first if k.startswith("_")]


def test_cases_json_keeps_the_request_block_verbatim(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="csv")
    cases = _cases_json(result)["cases"]
    with_request = [c for c in cases if c.get("request")]
    assert with_request, "样例用例里应当有带 request 块的"
    request = with_request[0]["request"]
    assert request["method"] == CASES[0]["request"]["method"]
    assert request["url"] == CASES[0]["request"]["url"]
    assert request["assertions"] == CASES[0]["request"]["assertions"]


def test_cases_json_omits_the_request_key_for_manual_cases(out_root):
    result = delivery.save_delivery([CASES[1]], title="人执行", fmt="csv")
    assert "request" not in _cases_json(result)["cases"][0]


def test_cases_json_round_trips_into_the_executor_shape(out_root):
    """落盘 → 读回 → 归一化，编号与请求块一条不少：工作台执行走的就是这条路。"""
    from server import case_validate

    result = delivery.save_delivery(CASES, title="回环", fmt="csv")
    back = case_validate.normalize_cases(_cases_json(result)["cases"])
    assert [c["case_id"] for c in back] == ["TC-001", "TC-002"]
    assert all(c["_shape_ok"] for c in back)
    assert back[0]["request"] == CASES[0]["request"]


def test_receipt_records_the_cases_file_hash(out_root):
    result = delivery.save_delivery(CASES, title="订单用例", fmt="csv")
    sidecar = _receipt(result)["cases_file"]
    assert sidecar["file"] == "cases.json"
    assert sidecar["schema"] == delivery.CASES_SCHEMA
    assert sidecar["case_count"] == 2
    blob = open(result["cases_file"], "rb").read()
    assert sidecar["sha256"] == hashlib.sha256(blob).hexdigest()
    assert sidecar["bytes"] == len(blob)


def test_cases_json_is_not_counted_as_a_format_artifact(out_root):
    """`artifacts` 的语义是"按 format 选出来的产物"，边车文件不混进去。"""
    result = delivery.save_delivery(CASES, title="订单用例", fmt="csv")
    assert [a["file"] for a in _receipt(result)["artifacts"]] == ["cases.csv"]


# ── login_request：批次自带"怎么重登录"（0.8，schema v2） ────────────────────
# 带上这份描述，用户在工作台上勾一下「登录换新」就能跑；不带就跟 0.7.0 一样。

LOGIN_SPEC = {
    "method": "POST",
    "url": "{{baseUrl}}/api/v1/auth/login",
    "path": "/api/v1/auth/login",
    "headers": [{"key": "Content-Type", "value": "application/json"}],
    "body": {"mode": "raw", "language": "json",
             "raw": '{"username": "{{login_username}}", '
                    '"password": "{{login_password}}"}'},
    "token_extract": {"source": "json_body", "path": "$.data.token"},
}


def test_login_request_is_stored_in_cases_json(out_root):
    result = delivery.save_delivery(CASES, title="带登录", fmt="csv",
                                    login_request=LOGIN_SPEC)
    payload = _cases_json(result)
    assert payload["login_request"] == LOGIN_SPEC
    assert "login_request_note" in payload
    assert _receipt(result)["cases_file"]["login_request"] is True


def test_without_login_request_the_key_is_absent(out_root):
    """不给就不写这个键——0.7.0 的批次形状一个字节都不变。"""
    result = delivery.save_delivery(CASES, title="不带登录", fmt="csv")
    payload = _cases_json(result)
    assert "login_request" not in payload
    assert _receipt(result)["cases_file"]["login_request"] is False


def test_login_request_accepts_a_json_string_and_a_single_key_wrapper(out_root):
    for value in (json.dumps(LOGIN_SPEC), {"login_request": LOGIN_SPEC}):
        result = delivery.save_delivery(CASES, title="宽容", fmt="csv",
                                        login_request=value)
        assert _cases_json(result)["login_request"] == LOGIN_SPEC


def test_a_junk_login_request_is_dropped_with_a_note(out_root):
    """救不回来就不写进落盘文件，但要如实回显——静默丢弃是最坏的一种。"""
    result = delivery.save_delivery(CASES, title="坏描述", fmt="csv",
                                    login_request="这不是 JSON")
    assert result["ok"] is True
    assert "login_request" not in _cases_json(result)
    assert any("login_request" in n for n in result["normalized"])


# ── 导出侧 PII 脱敏（BB-424）────────────────────────────────────────────────
#
# BB-424 的暴露面不是"本地留了 PII"，而是"**拿去分享的产物**里带着真人信息"。
# 所以闸开在导出产物上，cases.json 不动——那份是本地执行用的边车，
# 值被换成 <手机号> 会让执行真把这六个字符发出去。
# 两份文件的分工是刻意的，下面两条测试正反面各钉一次。

PII_CASE = [{
    "编号": "TC-001",
    "标题": "下单成功",
    "前置条件": "已登录",
    "操作步骤": ["调用 POST /api/order"],
    "预期结果": "状态码 200",
    "优先级": "高",
    "所属模块": "订单",
    "测试数据": "收件人张三，手机 13900139000，身份证 440305199001011234",
    "request": {
        "method": "POST",
        "url": "https://api.shop.example.com/api/order",
        "headers": [{"key": "Authorization", "value": "{{token}}"}],
        "body": {"mode": "raw", "language": "json",
                 "raw": '{"receiver": "张三", "mobile": "13900139000"}'},
        "assertions": [{"type": "status", "expected": 200}],
    },
}]

_PII_VALUES = ("13900139000", "440305199001011234", "张三")


def _read(path):
    return open(path, encoding="utf-8-sig", errors="ignore").read()


def test_export_products_carry_no_raw_pii(tmp_path):
    """有形态特征的几类（手机号/身份证/邮箱/卡号）必须从产物里消失。"""
    result = delivery.save_delivery(PII_CASE, fmt="csv", title="pii",
                                    out_root=str(tmp_path))
    assert result["ok"] is True
    csv_text = _read(os.path.join(result["delivery_dir"], "cases.csv"))
    for real in ("13900139000", "440305199001011234"):
        assert real not in csv_text, f"{real} 留在了要分享出去的 csv 里"
    assert "<手机号>" in csv_text, "脱敏要保形，不能换成 *** 让用例失去含义"


def test_a_name_in_free_text_is_a_known_gap_not_a_silent_one(tmp_path):
    """**这条钉的是局限，不是能力。**

    中文姓名没有形态特征（「张三」和「首页」都是两个汉字），只能靠键名缩小范围。
    所以结构化字段 `{"receiver": "张三"}` 里的姓名抓得到，而
    「收件人张三，手机…」这种自由文本里的抓不到。

    把它写成测试而不是留在文档里，是因为：日后有人看到产物里还有名字，
    要能一眼查到这是**已知边界**而非新 bug；反过来若哪天上了 NER 补上了这一类，
    这条会转红，那时该改的是断言不是删闸。
    收据里的 note 必须同步说明这一点——不许让用户以为产物已经干净了。
    """
    result = delivery.save_delivery(PII_CASE, fmt="csv", title="pii",
                                    out_root=str(tmp_path))
    csv_text = _read(os.path.join(result["delivery_dir"], "cases.csv"))
    assert "张三" in csv_text, "自由文本姓名如果被抓到了，说明能力提升了，请更新本测试"
    note = result["receipt"]["pii_redaction"]["note"]
    assert "自由文本" in note, "局限必须写在留痕里，否则用户会以为产物已无个人信息"


def test_the_name_in_a_structured_field_is_scrubbed(tmp_path):
    """对照面：键名像姓名时（body 里的 receiver）就抓得到。"""
    result = delivery.save_delivery(PII_CASE, fmt="postman", title="pii",
                                    out_root=str(tmp_path))
    blob = _read(result["postman_file"])
    assert '"receiver": "张三"' not in blob and '"receiver":"张三"' not in blob
    assert "<姓名>" in blob


def test_postman_collection_is_scrubbed_too(tmp_path):
    """collection 最容易被整份发给别人，它必须和表格一样干净。"""
    result = delivery.save_delivery(PII_CASE, fmt="postman", title="pii",
                                    out_root=str(tmp_path))
    blob = _read(result["postman_file"])
    for real in ("13900139000", "440305199001011234"):
        assert real not in blob


def test_cases_json_keeps_the_original_values_for_execution(tmp_path):
    """反面：边车不脱敏。改动这条前先想清楚执行层要拿什么去发请求。"""
    result = delivery.save_delivery(PII_CASE, fmt="csv", title="pii",
                                    out_root=str(tmp_path))
    payload = json.loads(_read(result["cases_file"]))
    blob = json.dumps(payload, ensure_ascii=False)
    assert "13900139000" in blob, "cases.json 被脱敏了，执行时会把占位符当真值发出去"


def test_receipt_records_the_redaction_even_when_nothing_was_hit(tmp_path):
    """留痕无条件写。

    只在命中时才写，会让"没这段"同时意味着"没 PII"和"闸没开"——读的人分不清。
    """
    clean = [dict(PII_CASE[0], 测试数据="page=1",
                  request=dict(PII_CASE[0]["request"], body={"mode": "none"}))]
    result = delivery.save_delivery(clean, fmt="csv", title="clean",
                                    out_root=str(tmp_path))
    section = result["receipt"]["pii_redaction"]
    assert section["applied"] is True
    assert section["hits"] == {}


def test_redaction_hits_are_reported_not_silent(tmp_path):
    """替换了什么、几处，必须说得出。"""
    result = delivery.save_delivery(PII_CASE, fmt="csv", title="pii",
                                    out_root=str(tmp_path))
    hits = result["pii_redaction"]["hits"]
    assert hits.get("手机号", 0) >= 1 and hits.get("身份证", 0) >= 1
    assert hits.get("姓名", 0) >= 1, "BB-465：序列化 body 里的姓名也要算进来"


def test_redaction_can_be_turned_off_explicitly(tmp_path):
    """留一个显式关闭口：有人确实需要带真实测试数据的产物。

    默认是开——安全默认不该要求用户先知道有这么个开关。
    """
    result = delivery.save_delivery(PII_CASE, fmt="csv", title="pii",
                                    out_root=str(tmp_path), redact_pii=False)
    csv_text = _read(os.path.join(result["delivery_dir"], "cases.csv"))
    assert "13900139000" in csv_text
    assert result["receipt"]["pii_redaction"]["applied"] is False
