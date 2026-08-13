"""交付落盘（`save_delivery` 工具的实现层）。

落盘目录：`deliveries/<YYYYMMDD-HHmmss>-<title-slug>/`
产物：`cases.xlsx` / `cases.csv` / `cases.md` / `<slug>.postman_collection.json`
（按 format 选，可多产物）+ `cases.json` + `receipt.json`。

## `cases.json` 是给机器读的那一份（0.7.0 起，无条件落盘）

xlsx 是给人读的、collection 是给 Postman 跑的，两者都不是"把这批用例原样读回来"
的好载体（xlsx 丢了 `request` 块，collection 把断言编译成了 pm.test 的 JS 文本，
反解回结构等于自己写一个 JS 解释器）。工作台要做的事是「列出这批用例 → 勾几条 →
交给 `execute_cases` 跑」，它需要的正是**原始用例数组**，所以直接落一份。

它不进 `receipt.json` 的 `artifacts`：那个清单的语义是"按 format 选出来的交付产物"，
而 `cases.json` 与 `receipt.json` 一样是随每次交付无条件落盘的边车文件。
它的哈希与大小记在收据的 `cases_file` 里，同样可对账。

### `login_request`：让批次自带"怎么重登录"（0.8，schema v2）

HAR 里录的 token 会过期。`parse_har` 对档A HAR 会产出 `replay.login_request`
（登录请求的可执行描述，账密位置是 `{{login_username}}` / `{{login_password}}` 占位，
**真值不在里面**）。落盘时把它一并存进 `cases.json`，批次就自带"怎么换新 token"，
工作台上勾一下「登录换新」就能跑——不必回聊天窗口把那个对象再传一遍。
不给这个参数就不写这个键，`cases.json` 与 0.7.0 那份完全一致。

M2.5「用例可执行化」：format 增加 `postman`（Collection v2.1，Apifox 原生可导入）
与 `xlsx+postman` 双产物，后者是 HAR 链路的新默认——xlsx 给人读与进 TAPD，
collection 给机器跑。翻译逻辑在 `server/postman.py`。

列名对齐 TAPD 用例库（PLAN M2 字段分层的「落盘全字段」）：
    用例名称 / 所属模块 / 前置条件 / 用例步骤 / 预期结果 / 用例等级 / 用例类型 /
    测试数据 / 关联端点
用例「编号」不是 TAPD 导入列，不进表格，改记进 `receipt.json` 的 `case_index`
（编号→用例名称），保证校验报错能回溯到落盘条目。Markdown 产物是给人评审的，
额外多一列编号。

轻量治理（PLAN M2）：收据只记输入指纹、生成时间、校验结果摘要、产物清单与各自
sha256——无哈希链、无装箱关。
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import time
from typing import Any

from server import args_tolerance, case_validate, postman
# 导出侧的 PII 脱敏复用出境闸那一套（BB-424 修复）。方向安全：`server/generate/`
# 不反向 import 本模块，不成环。刻意不另造一套词表——BB-424 的成因就是
# "第二套脱敏规则漏了一整类"，再造第三套只会重演。
from server.generate.scrub import scrub_payload

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DELIVERIES_DIR = os.path.join(REPO_ROOT, "deliveries")

RECEIPT_SCHEMA = "test-partner.delivery-receipt/v1"
#: `cases.json` 的 schema 标识（工作台按它认这份文件；改结构就升版本号）
CASES_SCHEMA = "test-partner.delivery-cases/v2"
CASES_FILE = "cases.json"
FORMATS = ("xlsx", "csv", "markdown", "postman", "xlsx+postman")
#: HAR 链路默认双产物（M2.5 用户拍板：xlsx 给人读/进 TAPD，collection 给机器跑）
DEFAULT_FORMAT = "xlsx+postman"
#: format 别名 → 规范值
FORMAT_ALIASES = {
    "md": "markdown", "markdown": "markdown",
    "excel": "xlsx", "xls": "xlsx", "xlsx": "xlsx",
    "csv": "csv",
    "postman": "postman", "collection": "postman",
    "postman_collection": "postman", "apifox": "postman",
    "xlsx+postman": "xlsx+postman", "postman+xlsx": "xlsx+postman",
    "excel+postman": "xlsx+postman", "both": "xlsx+postman",
}

#: TAPD 用例库列名（顺序即导出列顺序）
TAPD_COLUMNS = ("用例名称", "所属模块", "前置条件", "用例步骤", "预期结果",
                "用例等级", "用例类型", "测试数据", "关联端点")
DEFAULT_CASE_TYPE = "功能测试"
DEFAULT_MODULE = "未分类"

_SLUG_KEEP_RE = re.compile(r"[^0-9A-Za-z\u4e00-\u9fff]+")


class DeliveryError(ValueError):
    """落盘层可预期的错误（格式不支持、目录写不了等）。"""

    def __init__(self, code: str, message: str, hint: str = ""):
        self.code = code
        self.message = message
        self.hint = hint
        super().__init__(message)


def slugify(title: str) -> str:
    slug = _SLUG_KEEP_RE.sub("-", str(title or "").strip()).strip("-")
    return (slug[:40] or "untitled")


def _steps_text(steps: list) -> str:
    if not steps:
        return ""
    if len(steps) == 1:
        return steps[0]
    return "\n".join(f"{i}. {s}" for i, s in enumerate(steps, 1))


def to_rows(cases: Any) -> tuple[list, list]:
    """归一化用例 →（TAPD 列的行字典列表, 编号索引）。"""
    rows: list = []
    index: list = []
    for case in case_validate.normalize_cases(cases):
        if not case["_shape_ok"]:
            continue
        title = case["title"] or "(无标题)"
        rows.append({
            "用例名称": title,
            "所属模块": case["module"] or DEFAULT_MODULE,
            "前置条件": case["preconditions"] or "无",
            "用例步骤": _steps_text(case["steps"]),
            "预期结果": case["expected"],
            "用例等级": case["priority"],
            "用例类型": case["case_type"] or DEFAULT_CASE_TYPE,
            "测试数据": case["test_data"],
            "关联端点": "; ".join(case["endpoints"]),
        })
        index.append({"编号": case["case_id"], "用例名称": title})
    return rows, index


#: `cases.json` 里每条用例的字段顺序（与 `case_validate.FIELD_ALIASES` 同名同义）
CASE_RECORD_FIELDS = ("case_id", "title", "module", "priority", "case_type",
                      "preconditions", "steps", "expected", "test_data",
                      "endpoints", "request")


def to_case_records(cases: Any) -> list:
    """归一化用例 → 结构化记录列表（`cases.json` 的 `cases` 段）。

    只留归一化后的公共字段，`_shape_ok` / `_index` / `_present` 这些内部标记不落盘
    （`_present` 还是个 `set`，本来也序列化不了）。`request` 块原样保留——
    它是工作台把这批用例交回 `execute_cases` 时唯一有用的东西。
    """
    out: list = []
    for case in case_validate.normalize_cases(cases):
        if not case["_shape_ok"]:
            continue
        row = {field: case.get(field) for field in CASE_RECORD_FIELDS}
        row["title"] = row["title"] or "(无标题)"
        if row.get("request") is None:
            row.pop("request")          # 人执行用例：没有请求块就不写这个键
        out.append(row)
    return out


def scrub_cases_for_export(cases: Any) -> tuple[list, dict]:
    """导出产物专用：把用例里的个人信息换成保形占位符（BB-424）。

    **只作用于要交出去的产物**（xlsx / csv / markdown / postman），
    不碰 `cases.json`——那份是本地执行用的边车，值被换掉会让执行发出
    `<手机号>` 这种字面量。两份的分工是刻意的：

    | 文件 | 谁看 | 脱敏 |
    |---|---|---|
    | cases.json | 只有本人，执行层读它 | 否 |
    | xlsx/csv/md/postman | **导出给别人**（评审、导入 Postman、贴文档） | 是 |

    这正是 BB-424 的暴露面：产物是拿来分享的，而分享出去的东西里不该带
    真实身份证号。脱敏是**保形**的（`13800138000` → `<手机号>`），
    用例作为文档反而更正确——测试用例本来就不该硬编码真人信息。

    返回 `(脱敏后的用例, {类型: 命中数})`。命中数要一路回传到界面：
    静默替换会让用户以为产物里还是原值，拿去执行时才发现对不上。
    """
    cleaned, hits = scrub_payload(list(cases) if isinstance(cases, (list, tuple)) else cases)
    return (cleaned if isinstance(cleaned, list) else list(cases)), hits


def coerce_login_request(value: Any) -> dict | None:
    """`login_request` 入参规整：JSON 字符串解开；不是非空对象就当没给（返回 None）。

    形状对不对由执行层的闸说了算（那里报错才能说清后果），这里只负责"别把垃圾写进
    落盘文件"。**不做任何脱敏**——parse_har 产出的这个对象里账密本来就是占位符。
    """
    if isinstance(value, str) and value.strip():
        try:
            value = json.loads(value)
        except ValueError:
            return None
    if isinstance(value, dict) and len(value) == 1:
        only = next(iter(value))
        if str(only).strip().lower() in ("login_request", "loginrequest", "login"):
            value = value[only]
    return value if isinstance(value, dict) and value else None


def _write_cases_json(path: str, records: list, title: str, generated_at: str,
                      source_fingerprint: str, login_request: Any = None) -> None:
    payload = {
        "schema": CASES_SCHEMA,
        "title": title,
        "generated_at": generated_at,
        "source_fingerprint": source_fingerprint or None,
        "case_count": len(records),
        "note": ("给机器读的那一份：工作台按它列用例、勾选后交给 execute_cases 执行。"
                 "凭据值不在这里——请求块里的凭证位置写的是 {{变量名}}，"
                 "值在本机配置页的测试环境里。"),
        "cases": records,
    }
    if login_request:
        payload["login_request"] = login_request
        payload["login_request_note"] = (
            "登录换新用的登录请求描述（来自 parse_har 的 replay.login_request）。"
            "账号口令是 {{login_username}} / {{login_password}} 占位，真值在配置页的"
            "测试环境里。工作台选「登录换新」时就是把它交给 execute_cases。")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def _sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _write_csv(path: str, rows: list) -> None:
    # utf-8-sig：Excel 双击打开不乱码
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(TAPD_COLUMNS))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_markdown(path: str, rows: list, index: list, title: str,
                    generated_at: str, source_fingerprint: str) -> None:
    def cell(text: str) -> str:
        return str(text or "").replace("|", "\\|").replace("\n", "<br>")

    lines = [f"# {title}", "",
             f"- 生成时间：{generated_at}",
             f"- 用例条数：{len(rows)}",
             f"- 来源指纹：{source_fingerprint or '（未提供）'}", "",
             "| 编号 | " + " | ".join(TAPD_COLUMNS) + " |",
             "| --- | " + " | ".join("---" for _ in TAPD_COLUMNS) + " |"]
    for row, idx in zip(rows, index):
        cells = [cell(idx["编号"])] + [cell(row[c]) for c in TAPD_COLUMNS]
        lines.append("| " + " | ".join(cells) + " |")
    lines.append("")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines))


def _write_xlsx(path: str, rows: list) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:      # pragma: no cover - 依赖缺失时的可读错误
        raise DeliveryError(
            "XLSX_DEPENDENCY_MISSING",
            "缺少 openpyxl，无法产出 Excel",
            "在仓库根跑 .venv\\Scripts\\python.exe -m pip install -r requirements.txt，"
            "或改用 format=\"csv\"。") from exc

    widths = {"用例名称": 38, "所属模块": 14, "前置条件": 26, "用例步骤": 46,
              "预期结果": 40, "用例等级": 10, "用例类型": 12, "测试数据": 26,
              "关联端点": 30}
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(list(TAPD_COLUMNS))
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    for col, name in enumerate(TAPD_COLUMNS, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = widths[name]
    for row in rows:
        ws.append([row[c] for c in TAPD_COLUMNS])
    wrap = Alignment(wrap_text=True, vertical="top")
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(TAPD_COLUMNS) + 1):
            ws.cell(row=r, column=c).alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(path)


def _write_postman(path: str, cases: Any, title: str, source_fingerprint: str) -> dict:
    collection, stats = postman.build_collection(
        cases, title=title, source_fingerprint=source_fingerprint)
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(collection, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return stats


def normalize_format(fmt: str) -> str:
    """format 入参 → 规范值；不认识的抛 DeliveryError。"""
    text = str(fmt or DEFAULT_FORMAT).strip().strip("\"'").lower().replace(" ", "")
    normalized = FORMAT_ALIASES.get(text)
    if normalized is None:                  # 再宽容一档：别的分隔符按「+」再认一次
        normalized = FORMAT_ALIASES.get(re.sub(r"[-_,&/]+", "+", text))
    if normalized is None:
        raise DeliveryError("FORMAT_UNSUPPORTED", f"不支持的格式：{fmt}",
                            f"只支持 {'/'.join(FORMATS)}。")
    return normalized


def save_delivery(cases: Any, fmt: str = DEFAULT_FORMAT, title: str = "",
                  source_fingerprint: str = "", login_request: Any = None,
                  out_root: str = "", redact_pii: bool = True) -> dict:
    """落盘交付产物 + 收据。出错返回带 error 字段的可读结果，不抛裸异常。

    `out_root` 留空时落在仓库根的 `deliveries/`（MCP 工具那条线的既有行为，
    一行不改）。工作台走 HTTP 面时会把**当前用户的**批次目录传进来——
    决策 0009 要求按用户隔离，而 `DELIVERIES_DIR` 是模块常量、全进程共用一个。

    为什么是参数而不是请求期改写模块常量：同进程里两个用户同时采纳会互踩，
    改常量这种进程级可变状态在并发下必然出错。

    入参形状先过 `args_tolerance`（cases 的 JSON 字符串/单键包裹/单个用例对象、
    format 的大小写与空格都救回来），救回的动作如实回显在返回值的 `normalized` 里。
    """
    notes: list = []
    try:
        cases, fmt, title, source_fingerprint, notes = \
            args_tolerance.tolerant_delivery_args(cases, fmt, title, source_fingerprint)
        fmt = normalize_format(fmt)
        login_plan = coerce_login_request(login_request)
        if login_request is not None and login_plan is None:
            notes.append("login_request 不是可用的对象（也不是合法 JSON 字符串），"
                         "已忽略——这批用例的 cases.json 里不带登录换新描述")
        if not isinstance(cases, (list, tuple)) or not cases:
            raise DeliveryError("CASES_EMPTY", "cases 必须是非空的用例数组")

        # 导出产物走脱敏副本，cases.json 仍用原文（见 scrub_cases_for_export）。
        export_cases, pii_hits = (scrub_cases_for_export(cases) if redact_pii
                                  else (list(cases), {}))
        rows, index = to_rows(export_cases)
        if not rows:
            raise DeliveryError("CASES_ALL_INVALID",
                                "没有任何一条用例是合法对象，无可落盘内容")

        title = str(title or "").strip() or "测试用例"
        slug = slugify(title)
        stamp = time.strftime("%Y%m%d-%H%M%S")
        base_dir = os.path.join(out_root or DELIVERIES_DIR, f"{stamp}-{slug}")
        out_dir, bump = base_dir, 1
        while os.path.exists(out_dir):      # 同秒同题的第二次交付不覆盖前一次
            bump += 1
            out_dir = f"{base_dir}-{bump}"
        os.makedirs(out_dir)

        generated_at = time.strftime("%Y-%m-%d %H:%M:%S")
        filenames = {"xlsx": "cases.xlsx", "csv": "cases.csv", "markdown": "cases.md",
                     "postman": f"{slug}.postman_collection.json"}
        products: list = []
        postman_stats: dict = {}
        postman_path = ""
        for one in fmt.split("+"):
            product = os.path.join(out_dir, filenames[one])
            if one == "xlsx":
                _write_xlsx(product, rows)
            elif one == "csv":
                _write_csv(product, rows)
            elif one == "markdown":
                _write_markdown(product, rows, index, title, generated_at,
                                source_fingerprint)
            else:
                postman_stats = _write_postman(product, export_cases, title,
                                               source_fingerprint)
                postman_path = product
            products.append(product)

        # 边车文件：无条件落一份结构化用例，工作台按它列表与执行（见模块 docstring）
        case_records = to_case_records(cases)
        cases_json_path = os.path.join(out_dir, CASES_FILE)
        _write_cases_json(cases_json_path, case_records, title, generated_at,
                          source_fingerprint, login_plan)

        warnings: list = []
        if postman_stats.get("placeholder_count"):
            warnings.append(
                f"{postman_stats['placeholder_count']}/{postman_stats['item_count']} "
                f"条用例没有 request 块，collection 里是占位 item（{postman.PLACEHOLDER_NOTE}）")
        if postman_stats.get("items_without_test"):
            warnings.append(
                f"{postman_stats['items_without_test']} 条用例没有可执行断言，"
                "导入后跑完无从判定成败（建议给 request.assertions 补断言）")
        if postman_stats.get("skipped_assertions"):
            warnings.append(
                f"{postman_stats['skipped_assertions']} 条断言翻不成 pm.test，已跳过"
                "（先跑 validate_cases 看 E15）")
        if postman_stats and not postman_stats.get("base_url"):
            warnings.append("collection 的 baseUrl 变量为空，导入 Apifox/Postman 后需自行填环境地址")

        validation = case_validate.validate_cases(cases)
        input_material = json.dumps(list(cases), ensure_ascii=False, sort_keys=True,
                                    separators=(",", ":")).encode("utf-8")
        receipt = {
            "schema": RECEIPT_SCHEMA,
            "title": title,
            "generated_at": generated_at,
            "format": fmt,
            "case_count": len(rows),
            "source_fingerprint": source_fingerprint or None,
            "input_fingerprint": "sha256:" + hashlib.sha256(input_material).hexdigest(),
            "validation": {
                "ok": validation["ok"],
                "error_count": validation["summary"]["error_count"],
                "warning_count": validation["summary"]["warning_count"],
                "verdict": validation["summary"]["verdict"],
                "errors": validation["errors"][:10],
                "warnings": validation["warnings"][:10],
                "note": "落盘时按 validate_cases 同一套规则复校（未传 endpoints，不含覆盖率）。",
            },
            "columns": list(TAPD_COLUMNS),
            "case_index": index,
            "artifacts": [{
                "file": os.path.basename(p),
                "sha256": _sha256_file(p),
                "bytes": os.path.getsize(p),
            } for p in products],
            # 边车文件单列一项：语义上不是"按 format 选的产物"，但一样要能对账
            "cases_file": {
                "file": CASES_FILE,
                "schema": CASES_SCHEMA,
                "case_count": len(case_records),
                "sha256": _sha256_file(cases_json_path),
                "bytes": os.path.getsize(cases_json_path),
                # 只记"带没带"这个事实，登录端点与字段名都在 cases.json 里，不重复
                "login_request": bool(login_plan),
            },
            # 导出侧脱敏的留痕（BB-424）。**无论有没有命中都写这一段**——
            # 只在命中时才写，会让"没这段"同时意味着"没 PII"和"闸没开"，读的人分不清。
            "pii_redaction": {
                "applied": bool(redact_pii),
                "scope": "导出产物（xlsx/csv/markdown/postman）；cases.json 保留原值供本地执行",
                "hits": pii_hits,
                "note": ("按形态识别的个人信息已换成保形占位符：身份证、手机号、邮箱、"
                         "银行卡、IP、长标识，以及**键名像姓名的字段**里的中文姓名。"
                         "抓不到的：自由文本里的姓名（「收件人张三，手机…」这种句子——"
                         "中文姓名没有形态特征，只能靠键名缩小范围）、住址、生日、护照号、车牌。"
                         "**所以这不等于产物已无个人信息，对外发之前仍请自行过一眼。**"
                         if redact_pii else
                         "本次导出未做个人信息脱敏（调用方显式关闭）。"),
            },
            "generator": {"server": "test-partner", "tool": "save_delivery"},
            "governance": "轻量治理：无哈希链、无装箱关；收据只记指纹/时间/校验摘要/产物哈希。",
        }
        if postman_stats:
            receipt["postman"] = dict(postman_stats,
                                      schema=postman.SCHEMA_URL,
                                      file=os.path.basename(postman_path))
        if notes:
            receipt["normalized"] = list(notes)
        if warnings:
            receipt["warnings"] = warnings
        receipt_path = os.path.join(out_dir, "receipt.json")
        with open(receipt_path, "w", encoding="utf-8", newline="\n") as f:
            json.dump(receipt, f, ensure_ascii=False, indent=2)
            f.write("\n")

        result = {
            "ok": True,
            "schema": RECEIPT_SCHEMA,
            "delivery_dir": out_dir,
            "files": products + [cases_json_path, receipt_path],
            "cases_file": cases_json_path,
            "format": fmt,
            "case_count": len(rows),
            "validation_ok": validation["ok"],
            "pii_redaction": receipt["pii_redaction"],
            "receipt": receipt,
        }
        if postman_path:
            result["postman_file"] = postman_path
            result["postman"] = receipt["postman"]
            result["import_hint"] = (
                f"{os.path.basename(postman_path)} 是标准 Postman Collection v2.1，"
                "可直接导入 Apifox（导入 → Postman）或 Postman 执行。")
        if warnings:
            result["warnings"] = warnings
        if notes:
            result["normalized"] = notes
        return result
    except args_tolerance.ArgsToleranceError as exc:
        return {"ok": False, "error": exc.code, "message": exc.message,
                "hint": exc.hint, "normalized": exc.normalized,
                "deliveries_dir": out_root or DELIVERIES_DIR}
    except DeliveryError as exc:
        result = {"ok": False, "error": exc.code, "message": exc.message,
                  "hint": exc.hint, "deliveries_dir": out_root or DELIVERIES_DIR}
        if notes:
            result["normalized"] = notes
        return result
    except Exception as exc:  # noqa: BLE001 - 工具边界收口
        return {"ok": False, "error": "DELIVERY_FAILED",
                "message": f"落盘出错：{type(exc).__name__}: {exc}",
                "hint": "确认交付目录可写、产物文件未被 Excel 占用。",
                "deliveries_dir": out_root or DELIVERIES_DIR}
